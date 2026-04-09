#!/usr/bin/env python3
"""
Excel Export Utility for Financial Analysis Results
Exports retained earnings and reinvested earnings data to formatted Excel files
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from datetime import datetime
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, output_dir: str = "output/excel"):
        # Use absolute path for output directory
        project_root = Path(__file__).parent.parent.parent
        self.output_dir = project_root / output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(
            start_color="1e6641", end_color="1e6641", fill_type="solid"
        )
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.right_alignment = Alignment(horizontal="right", vertical="center")

    def export_dashboard_table(self, data):
        """
        Export only the dashboard table data in a simple format
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Data"  # Shorter title to avoid Excel issues

            # Get headers dynamically from the data columns and reverse for RTL layout
            headers = list(data.columns)[::-1]  # Reverse the order for RTL layout

            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.right_alignment  # Right-align headers for RTL
                cell.border = self.border

            # Add data rows
            for row_idx, (_, row) in enumerate(data.iterrows(), 2):
                row_data = []
                for col, header in enumerate(headers, 1):
                    value = row.get(header, "")

                    # Clean and format numeric values
                    if (
                        value
                        and str(value).strip()
                        and str(value).lower()
                        not in ["", "null", "undefined", "nan", "لايوجد"]
                    ):
                        try:
                            # Convert to float and format if it's a number
                            num_value = float(str(value).replace(",", ""))
                            if num_value != 0:  # Only format non-zero numbers
                                formatted_value = f"{num_value:,.0f}"
                            else:
                                formatted_value = "0"
                        except (ValueError, TypeError):
                            formatted_value = str(value)
                    else:
                        formatted_value = "لايوجد"

                    row_data.append(formatted_value)

                # Add row data with reversed order for RTL layout
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = self.data_font
                    cell.border = self.border
                    cell.alignment = self.right_alignment

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, ValueError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"financial_analysis_{timestamp}.xlsx"
            output_path = self.output_dir / filename

            # Save workbook with error handling
            try:
                wb.save(str(output_path))
                logger.info(f"Dashboard table exported: {output_path}")
                return str(output_path)
            except Exception as save_error:
                logger.error(f"Error saving Excel file: {save_error}")
                return None

        except Exception as e:
            logger.error(f"Error exporting dashboard table: {e}")
            return None


def main():
    """Export current data to Excel"""
    try:
        # Load data
        csv_path = Path("data/results/reinvested_earnings_results.csv")
        if not csv_path.exists():
            print("Error: CSV file not found")
            return

        data = pd.read_csv(csv_path)
        print(f"Loaded {len(data)} records")

        # Export
        exporter = ExcelExporter()
        output_path = exporter.export_dashboard_table(data)

        if output_path:
            print(f"✅ Excel file created: {output_path}")
        else:
            print("❌ Failed to create Excel file")

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()
